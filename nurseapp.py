from __future__ import annotations
from datetime import datetime, date, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st
from icalendar import Calendar
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =====================
# 金额 & OASIS 规则
# =====================

AMOUNT_RULES = {
    "SOC": 100,
    "DC": 60,
    "RECERT": 70,
    "ROC": 70,
    "IV": 65,
    "FU": 65,
}

OASIS_CODES = {"SOC", "DC", "RECERT", "ROC"}

YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NO_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

CLEAN_COLS = ["Date","Time In","Time Out","Patient Name","Visit Type","Location","Amount","OASIS Required?"]

# =====================
# Visit Type 标准化
# =====================

def normalize_visit_type(vt: str | None) -> str | None:
    if not vt:
        return None

    s = vt.upper().replace("-", " ").strip()

    if "SOC" in s:
        return "SOC"
    if "DC" in s:
        return "DC"
    if "RECERT" in s:
        return "RECERT"
    if "ROC" in s:
        return "ROC"
    if "FOLLOW UP" in s or s in {"FU", "F/U"}:
        return "FU"
    if "IV" in s:
        return "IV"

    return s

def calculate_amount(vt: str | None) -> int:
    return AMOUNT_RULES.get(normalize_visit_type(vt), 0)

def oasis_required(vt: str | None) -> str:
    return "YES" if normalize_visit_type(vt) in OASIS_CODES else "NO"

# =====================
# 时间安全处理
# =====================

def to_date_time(dt):
    """返回 (YYYY-MM-DD, HH:MM:SS)；如果只有 date 就 time=None"""
    if dt is None:
        return None, None

    if isinstance(dt, datetime):
        return dt.date().isoformat(), dt.time().strftime("%H:%M:%S")

    if isinstance(dt, date):
        return dt.isoformat(), None

    return str(dt), None

def to_iso(dt):
    """给 Raw 表用：尽量输出可读的 ISO 字符串"""
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.isoformat(sep=" ")
    if isinstance(dt, date):
        return dt.isoformat()
    return str(dt)

# =====================
# Summary 拆解
# =====================

def split_summary(summary: str | None):
    if not summary:
        return None, None

    s = summary.replace("—", "-").replace("–", "-").strip()

    if " - " in s:
        left, right = s.split(" - ", 1)
        return left.strip(), right.strip()

    if "-" in s:
        left, right = s.split("-", 1)
        return left.strip(), right.strip()

    return s.strip(), None

# =====================
# 取消/删除事件识别
# =====================

def is_cancelled_event(ev) -> bool:
    """
    常见 ICS 写法：
    - VEVENT: STATUS:CANCELLED
    - 少数导出会把 summary 写成 cancel/cancelled 开头
    """
    status = ev.get("status")
    if status and str(status).upper() == "CANCELLED":
        return True

    summary = ev.get("summary")
    if summary:
        s = str(summary).strip().lower()
        if s.startswith("cancelled") or s.startswith("canceled") or s.startswith("cancel"):
            return True

    return False

def calendar_method_is_cancel(cal: Calendar) -> bool:
    """少数 ICS 会在日历层写 METHOD:CANCEL"""
    m = cal.get("method")
    return bool(m and str(m).upper() == "CANCEL")

# =====================
# ICS 解析（返回 RAW + CLEAN）
# =====================

def parse_ics(file_bytes: bytes, include_cancelled_in_clean: bool = False) -> tuple[pd.DataFrame, pd.DataFrame]:
    cal = Calendar.from_ical(file_bytes)
    rows_clean = []
    rows_raw = []

    cal_cancel_mode = calendar_method_is_cancel(cal)

    def safe_int(x, default=0):
        try:
            return int(x)
        except Exception:
            return default

    for ev in cal.walk("VEVENT"):
        summary = str(ev.get("summary")) if ev.get("summary") else None
        location = str(ev.get("location")) if ev.get("location") else None
        uid = str(ev.get("uid")) if ev.get("uid") else None
        description = str(ev.get("description")) if ev.get("description") else None

        status = str(ev.get("status")).upper() if ev.get("status") else None
        sequence = safe_int(ev.get("sequence"), 0)

        last_modified = ev.get("last-modified").dt if ev.get("last-modified") else None
        dtstamp = ev.get("dtstamp").dt if ev.get("dtstamp") else None
        recurrence_id = ev.get("recurrence-id").dt if ev.get("recurrence-id") else None

        start = ev.get("dtstart").dt if ev.get("dtstart") else None
        end   = ev.get("dtend").dt if ev.get("dtend") else None

        start_iso = to_iso(start)
        end_iso   = to_iso(end)

        d, tin = to_date_time(start)
        _, tout = to_date_time(end)

        patient, raw_type = split_summary(summary)
        vt_norm = normalize_visit_type(raw_type)

        cancelled = cal_cancel_mode or is_cancelled_event(ev) or (status == "CANCELLED")

        # ✅ 事件指纹：UID 不可靠时也能尽量稳定
        uid_fallback = uid if uid else f"NOUID|{(summary or '')}|{(location or '')}"
        rid_iso = to_iso(recurrence_id) if recurrence_id else ""
        event_key = f"{uid_fallback}|RID={rid_iso}|{start_iso}|{end_iso}|{(summary or '')}|{(location or '')}"

        # Raw 永远保留（便于你对照“删了又回来”的来源）
        rows_raw.append({
            "EVENT_KEY": event_key,
            "UID": uid,
            "RECURRENCE-ID": to_iso(recurrence_id) if recurrence_id else None,
            "SEQUENCE": sequence,
            "LAST-MODIFIED": to_iso(last_modified),
            "DTSTAMP": to_iso(dtstamp),
            "STATUS": status,
            "CANCELLED?": "YES" if cancelled else "NO",

            "Summary (Raw)": summary,
            "Patient Name (Parsed)": patient,
            "Visit Type (Raw)": raw_type,
            "DTSTART (Raw)": start_iso,
            "DTEND (Raw)": end_iso,
            "Date (Parsed)": d,
            "Time In (Parsed)": tin,
            "Time Out (Parsed)": tout,
            "Location (Raw)": location,
            "Description (Raw)": description,
        })

        # Clean 默认不纳入 cancelled（除非你勾选）
        if cancelled and (not include_cancelled_in_clean):
            continue

        rows_clean.append({
            "EVENT_KEY": event_key,
            "Date": d,
            "Time In": tin,
            "Time Out": tout,
            "Patient Name": patient,
            "Visit Type": vt_norm,
            "Location": location,
            "Amount": calculate_amount(vt_norm),
            "OASIS Required?": oasis_required(vt_norm),
        })

    df_raw = pd.DataFrame(rows_raw)
    df_clean = pd.DataFrame(rows_clean)

    # =====================
    # 去重策略（核心）
    # 1) Raw：优先保留“最新版本”（SEQUENCE 最大，其次 LAST-MODIFIED/DTSTAMP 更晚）
    # 2) 再去 EVENT_KEY 的完全重复
    # 3) Clean：跟随 Raw 保留下来的 EVENT_KEY
    # =====================

    if len(df_raw):
        df_raw["_lm"] = pd.to_datetime(df_raw["LAST-MODIFIED"], errors="coerce")
        df_raw["_ds"] = pd.to_datetime(df_raw["DTSTAMP"], errors="coerce")

        # 先排序，让“最新的版本”排在最后
        df_raw = df_raw.sort_values(
            by=["UID", "RECURRENCE-ID", "DTSTART (Raw)", "SEQUENCE", "_lm", "_ds"],
            na_position="first"
        )

        # 同 UID + RID + DTSTART：保留最后一条（最新）
        # 注意：UID 空的会堆一起，所以只对 UID 非空的做“版本”去重
        non_empty_uid = df_raw[df_raw["UID"].notna() & (df_raw["UID"].astype(str).str.strip() != "")].copy()
        empty_uid = df_raw[~(df_raw["UID"].notna() & (df_raw["UID"].astype(str).str.strip() != ""))].copy()

        if len(non_empty_uid):
            non_empty_uid = non_empty_uid.drop_duplicates(
                subset=["UID", "RECURRENCE-ID", "DTSTART (Raw)"],
                keep="last"
            )

        # 合并回来
        df_raw = pd.concat([non_empty_uid, empty_uid], ignore_index=True)

        # 再去 EVENT_KEY 的完全重复
        df_raw = df_raw.drop_duplicates(subset=["EVENT_KEY"], keep="last")

        df_raw = df_raw.drop(columns=["_lm","_ds"], errors="ignore")

        # Clean 只保留 Raw 最终存在的 key（避免多版本/重复污染 Clean）
        keep_keys = set(df_raw["EVENT_KEY"].dropna().tolist())
        if len(df_clean):
            df_clean = df_clean[df_clean["EVENT_KEY"].isin(keep_keys)].copy()

    return df_raw, df_clean

# =====================
# Excel 导出
# =====================

def export_excel_clean(df: pd.DataFrame, path: Path):
    df[CLEAN_COLS].to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active
    col = CLEAN_COLS.index("OASIS Required?") + 1

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, col)
        cell.fill = YES_FILL if cell.value == "YES" else NO_FILL

    wb.save(path)

def export_excel_raw(df_raw: pd.DataFrame, path: Path):
    df_raw.to_excel(path, index=False)

def export_excel_bundle(df_clean: pd.DataFrame, df_raw: pd.DataFrame, path: Path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_clean[CLEAN_COLS].to_excel(writer, sheet_name="Clean_Amount", index=False)
        df_raw.to_excel(writer, sheet_name="Raw_Original", index=False)

        ws = writer.book["Clean_Amount"]
        header = [c.value for c in ws[1]]
        oasis_col = header.index("OASIS Required?") + 1

        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, oasis_col)
            cell.fill = YES_FILL if cell.value == "YES" else NO_FILL

# =====================
# 重复检测报告
# =====================

def build_duplicate_report(df_raw: pd.DataFrame, df_clean: pd.DataFrame) -> dict:
    tables: dict[str, pd.DataFrame] = {}

    # UID 缺失
    if len(df_raw):
        miss_uid = df_raw[df_raw["UID"].isna() | (df_raw["UID"].astype(str).str.strip() == "")]
        tables["UID 缺失（Raw）"] = miss_uid.copy()

    # cancelled（Raw）
    if len(df_raw) and "CANCELLED?" in df_raw.columns:
        canc = df_raw[df_raw["CANCELLED?"] == "YES"].copy()
        tables["被取消/删除事件（Raw）"] = canc

    # 同 UID 多版本（Raw）：同 UID + RID + DTSTART 多条
    if len(df_raw):
        key_cols = ["UID", "RECURRENCE-ID", "DTSTART (Raw)"]
        tmp = df_raw[df_raw["UID"].notna() & (df_raw["UID"].astype(str).str.strip() != "")].copy()
        if len(tmp):
            vc = tmp.groupby(key_cols, dropna=False).size().reset_index(name="count")
            multi = vc[vc["count"] > 1]
            if len(multi):
                multi_ver = tmp.merge(multi[key_cols], on=key_cols, how="inner")
                multi_ver = multi_ver.sort_values(key_cols + ["SEQUENCE"], na_position="last")
                tables["同一 UID 的多版本（Raw）"] = multi_ver

    # EVENT_KEY 完全重复（Raw）
    if len(df_raw) and "EVENT_KEY" in df_raw.columns:
        dk = df_raw[df_raw.duplicated(subset=["EVENT_KEY"], keep=False)].copy()
        if len(dk):
            tables["EVENT_KEY 完全重复（Raw）"] = dk

    # Clean 疑似重复：同 Date + Time In + Patient Name 多条
    if len(df_clean):
        tmpc = df_clean.copy()
        tmpc["Date"] = pd.to_datetime(tmpc["Date"], errors="coerce").dt.date
        base_cols = ["Date", "Time In", "Patient Name"]
        if all(c in tmpc.columns for c in base_cols):
            ct = tmpc.groupby(base_cols, dropna=False).size().reset_index(name="count")
            ct = ct[ct["count"] > 1]
            if len(ct):
                susp = tmpc.merge(ct[base_cols], on=base_cols, how="inner").sort_values(base_cols)
                tables["同病人同时间重复（Clean，疑似）"] = susp

    # summary
    summary_rows = [{"Issue Type": k, "Rows": int(len(v))} for k, v in tables.items()]
    if not summary_rows:
        summary_rows = [{"Issue Type": "未检测到明显重复/异常", "Rows": 0}]
    summary = pd.DataFrame(summary_rows)

    return {"summary": summary, "tables": tables}

def export_dup_report_excel(report: dict, path: Path):
    summary = report["summary"]
    tables = report["tables"]

    def safe_sheet_name(name: str) -> str:
        bad = ['\\', '/', '*', '?', ':', '[', ']']
        for b in bad:
            name = name.replace(b, "_")
        return name[:31] if len(name) > 31 else name

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        for k, df_ in tables.items():
            sn = safe_sheet_name(k)
            if df_ is None or len(df_) == 0:
                pd.DataFrame([{"info": "no rows"}]).to_excel(writer, sheet_name=sn, index=False)
            else:
                df_.to_excel(writer, sheet_name=sn, index=False)

# =====================
# Streamlit UI
# =====================

st.set_page_config(page_title="ICS Visits 管理与金额统计", layout="wide")
st.title("ICS Visits 管理与金额统计")

# ---- Session State ----
if "data" not in st.session_state:
    st.session_state.data = pd.DataFrame()

if "raw" not in st.session_state:
    st.session_state.raw = pd.DataFrame()

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

# ---- 取消事件导入开关（默认关闭） ----
include_cancelled_in_clean = st.checkbox(
    "包含已取消/已删除事件到 Clean（一般不建议）",
    value=False
)

# ---- 重置 ----
if st.button("🔄 重置记录（清空数据与文件）"):
    st.session_state.data = pd.DataFrame()
    st.session_state.raw = pd.DataFrame()
    st.session_state.uploader_key += 1
    st.success("已重置")

# ---- 上传 ----
uploaded = st.file_uploader(
    "📤 上传 .ics 文件",
    type=["ics"],
    key=f"uploader_{st.session_state.uploader_key}"
)

if uploaded:
    df_raw_new, df_clean_new = parse_ics(
        uploaded.getvalue(),
        include_cancelled_in_clean=include_cancelled_in_clean
    )
    st.session_state.raw = df_raw_new
    st.session_state.data = df_clean_new

df_raw = st.session_state.raw
df = st.session_state.data

# ===== Raw 表 =====
st.subheader("🧾 原版表（Raw，对照用，不做标准化）")
if len(df_raw):
    st.dataframe(df_raw, use_container_width=True, hide_index=True)
else:
    st.info("先上传 .ics 文件，Raw 表会出现在这里。")

# ===== 重复检测报告 =====
st.subheader("🧠 重复/异常检测报告（对照用）")
if len(df_raw) or len(df):
    report = build_duplicate_report(df_raw, df)
    cA, cB = st.columns([2, 1])
    with cA:
        st.dataframe(report["summary"], use_container_width=True, hide_index=True)
    with cB:
        st.info("常见来源：同 UID 多版本、STATUS=CANCELLED（删了又回魂）、UID 缺失导致去重不稳定。")

    with st.expander("展开查看明细（每类问题一张表）", expanded=False):
        for k, tdf in report["tables"].items():
            st.markdown(f"### {k}（{len(tdf)} rows）")
            st.dataframe(tdf, use_container_width=True, hide_index=True)

    out_rep = Path("Duplicate_Report.xlsx")
    export_dup_report_excel(report, out_rep)
    with open(out_rep, "rb") as f:
        st.download_button(
            "⬇️ 下载 Duplicate_Report.xlsx（推荐）",
            f,
            file_name=out_rep.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("先上传 .ics 文件，我才能检测重复。")

# ===== Clean 功能区 =====
if len(df):
    df = df.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date

    st.subheader("📆 日期范围统计")

    today = date.today()
    min_date = df["Date"].min()
    max_date = df["Date"].max()

    c1, c2 = st.columns([1, 3])

    with c1:
        preset = st.radio(
            "快速选择",
            ["本日", "本周", "本月", "上月", "全部", "自定义"],
            index=1,
        )

    with c2:
        if preset == "本日":
            start = end = today
        elif preset == "本周":
            start = today - timedelta(days=today.weekday())
            end = today
        elif preset == "本月":
            start = today.replace(day=1)
            end = today
        elif preset == "上月":
            first_this_month = today.replace(day=1)
            end = first_this_month - timedelta(days=1)
            start = end.replace(day=1)
        elif preset == "全部":
            start, end = min_date, max_date
        else:
            start, end = st.date_input(
                "选择日期范围",
                [min_date, max_date],
                min_value=min_date,
                max_value=max_date,
            )

    filt = df[(df["Date"] >= start) & (df["Date"] <= end)].copy()

    # Raw 也跟着筛选（用于对照导出）
    filt_raw = df_raw.copy()
    if len(filt_raw) and "Date (Parsed)" in filt_raw.columns:
        filt_raw["Date (Parsed)"] = pd.to_datetime(filt_raw["Date (Parsed)"], errors="coerce").dt.date
        filt_raw = filt_raw[(filt_raw["Date (Parsed)"] >= start) & (filt_raw["Date (Parsed)"] <= end)].copy()

    m1, m2 = st.columns(2)
    m1.metric("💰 总金额", f"${int(filt['Amount'].sum())}")
    m2.metric("📊 Visits 数量", len(filt))

    st.subheader("📋 Visits 明细（Clean，标准化后）")
    st.dataframe(
        filt.sort_values(["Date", "Time In"], na_position="last"),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("⬇️ 导出下载")

    out_clean = Path("HH_Visits_With_Amount.xlsx")
    export_excel_clean(filt, out_clean)
    with open(out_clean, "rb") as f:
        st.download_button(
            "⬇️ 下载 Excel（Clean，带金额+上色）",
            f,
            file_name=out_clean.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    out_raw = Path("HH_Visits_Raw.xlsx")
    export_excel_raw(filt_raw if len(filt_raw) else df_raw, out_raw)
    with open(out_raw, "rb") as f:
        st.download_button(
            "⬇️ 下载 Excel（Raw 原版对照）",
            f,
            file_name=out_raw.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    out_bundle = Path("HH_Visits_Bundle.xlsx")
    export_excel_bundle(filt, (filt_raw if len(filt_raw) else df_raw), out_bundle)
    with open(out_bundle, "rb") as f:
        st.download_button(
            "⬇️ 下载 Excel（Bundle：Clean+Raw 两个Sheet）",
            f,
            file_name=out_bundle.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

else:
    st.warning("还没有解析后的 Clean 数据。上传 .ics 后会显示统计与导出。")
