from __future__ import annotations
from datetime import datetime, date, timedelta
from pathlib import Path
import json
import os
import sys

# 必须在 PyQt6 导入之前设置，否则打包后找不到 cocoa 插件
if getattr(sys, "frozen", False) and sys.platform == "darwin":
    _exe_dir = os.path.dirname(os.path.abspath(sys.executable))
    for _p in [
        os.path.join(getattr(sys, "_MEIPASS", ""), "PyQt6", "Qt6", "plugins", "platforms"),
        os.path.join(_exe_dir, "..", "Frameworks", "PyQt6", "Qt6", "plugins", "platforms"),
        os.path.join(_exe_dir, "PyQt6", "Qt6", "plugins", "platforms"),
        os.path.join(_exe_dir, "_internal", "PyQt6", "Qt6", "plugins", "platforms"),
    ]:
        _p = os.path.normpath(_p)
        if os.path.isdir(_p):
            os.environ["QT_QPA_PLATFORM_PLUGIN_PATH"] = _p
            os.environ["QT_PLUGIN_PATH"] = os.path.dirname(_p)
            break

import pandas as pd
from icalendar import Calendar
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QCheckBox, QTabWidget, QComboBox, QDateEdit, QGroupBox,
    QHeaderView, QMessageBox, QDialog, QDialogButtonBox, QAbstractItemView,
)
from PyQt6.QtCore import Qt, QDate, QSettings
from PyQt6.QtGui import QColor, QFont


# =====================
# 国际化（i18n）
# =====================

TRANSLATIONS = {
    "zh": {
        "window_title": "ICS Visits 管理与金额统计",
        "lang_label": "语言:",
        "lang_zh": "中文",
        "lang_en": "English",
        "btn_upload": "📤 上传 .ics 文件",
        "btn_reset": "🔄 重置",
        "chk_cancelled": "包含已取消事件到 Clean（一般不建议）",
        "tab_clean": "📋 Clean 数据",
        "tab_raw": "🧾 Raw 原版",
        "tab_dup": "🧠 重复检测",
        "grp_date": "日期范围筛选",
        "lbl_quick": "快速选择:",
        "lbl_from": "从:",
        "lbl_to": "到:",
        "preset_today": "本日",
        "preset_week": "本周",
        "preset_month": "本月",
        "preset_last_month": "上月",
        "preset_all": "全部",
        "preset_custom": "自定义",
        "lbl_amount": "💰 总金额: ${amount}",
        "lbl_visits": "📊 Visits 数量: {count}",
        "btn_export_clean": "⬇️ 导出 Clean Excel(带金额+上色)",
        "btn_export_bundle": "⬇️ 导出 Bundle(Clean + Raw 两个 Sheet)",
        "btn_export_raw": "⬇️ 导出 Raw Excel(原版对照)",
        "btn_export_dup": "⬇️ 导出 Duplicate Report Excel",
        "lbl_dup_summary": "问题汇总:",
        "lbl_dup_detail": "明细(每类问题一个标签页):",
        "status_no_data": "请上传 .ics 文件",
        "status_reset": "已重置，请上传 .ics 文件",
        "status_loaded": "已加载: {name}  |  Raw: {raw} 条  |  Clean: {clean} 条",
        "dlg_select_ics": "选择 .ics 文件",
        "dlg_ics_filter": "iCalendar Files (*.ics)",
        "dlg_save_clean": "保存 Clean Excel",
        "dlg_save_raw": "保存 Raw Excel",
        "dlg_save_bundle": "保存 Bundle Excel",
        "dlg_save_dup": "保存 Duplicate Report",
        "dlg_xlsx_filter": "Excel Files (*.xlsx)",
        "msg_parse_failed": "解析失败",
        "msg_no_data_title": "无数据",
        "msg_no_data_filt": "请先上传并筛选数据。",
        "msg_no_data_raw": "请先上传数据。",
        "msg_no_report_title": "无报告",
        "msg_no_report": "请先上传数据生成报告。",
        "msg_saved": "已保存",
        "col_issue_type": "问题类型",
        "col_rows": "条数",
        "sheet_summary": "Summary",
        "report_missing_uid": "UID 缺失(Raw)",
        "report_cancelled": "被取消/删除事件(Raw)",
        "report_multi_version": "同一 UID 的多版本(Raw)",
        "report_duplicate_event_key": "EVENT_KEY 完全重复(Raw)",
        "report_same_patient_time": "同病人同时间重复(Clean，疑似)",
        "report_none": "未检测到明显重复/异常",
        # 访视类型编辑器
        "btn_settings": "⚙️ 访视类型设置",
        "rules_dialog_title": "访视类型与价格",
        "rules_help": "在这里添加你常用的访视类型。导入日历时，程序会按从上到下的顺序匹配关键词。",
        "rules_col_code": "代号",
        "rules_col_keywords": "匹配关键词（逗号分隔）",
        "rules_col_amount": "金额 ($)",
        "rules_col_oasis": "需要 OASIS",
        "btn_rule_add": "+ 添加",
        "btn_rule_del": "- 删除",
        "btn_rule_up": "↑ 上移",
        "btn_rule_down": "↓ 下移",
        "btn_rule_reset": "恢复默认",
        "msg_rule_invalid": "代号不能为空（第 {row} 行）",
        "msg_rule_duplicate": "代号 \"{code}\" 重复",
        "msg_rule_bad_amount": "金额必须是数字（第 {row} 行）",
    },
    "en": {
        "window_title": "ICS Visits Manager",
        "lang_label": "Language:",
        "lang_zh": "中文",
        "lang_en": "English",
        "btn_upload": "📤 Upload .ics File",
        "btn_reset": "🔄 Reset",
        "chk_cancelled": "Include cancelled events in Clean (not recommended)",
        "tab_clean": "📋 Clean Data",
        "tab_raw": "🧾 Raw Original",
        "tab_dup": "🧠 Duplicates",
        "grp_date": "Date Range Filter",
        "lbl_quick": "Quick select:",
        "lbl_from": "From:",
        "lbl_to": "To:",
        "preset_today": "Today",
        "preset_week": "This week",
        "preset_month": "This month",
        "preset_last_month": "Last month",
        "preset_all": "All",
        "preset_custom": "Custom",
        "lbl_amount": "💰 Total: ${amount}",
        "lbl_visits": "📊 Visits: {count}",
        "btn_export_clean": "⬇️ Export Clean Excel (amounts + colors)",
        "btn_export_bundle": "⬇️ Export Bundle (Clean + Raw sheets)",
        "btn_export_raw": "⬇️ Export Raw Excel (original)",
        "btn_export_dup": "⬇️ Export Duplicate Report",
        "lbl_dup_summary": "Issue summary:",
        "lbl_dup_detail": "Details (one tab per issue type):",
        "status_no_data": "Please upload an .ics file",
        "status_reset": "Reset done — please upload an .ics file",
        "status_loaded": "Loaded: {name}  |  Raw: {raw} rows  |  Clean: {clean} rows",
        "dlg_select_ics": "Select .ics file",
        "dlg_ics_filter": "iCalendar Files (*.ics)",
        "dlg_save_clean": "Save Clean Excel",
        "dlg_save_raw": "Save Raw Excel",
        "dlg_save_bundle": "Save Bundle Excel",
        "dlg_save_dup": "Save Duplicate Report",
        "dlg_xlsx_filter": "Excel Files (*.xlsx)",
        "msg_parse_failed": "Parse failed",
        "msg_no_data_title": "No data",
        "msg_no_data_filt": "Please upload and filter data first.",
        "msg_no_data_raw": "Please upload data first.",
        "msg_no_report_title": "No report",
        "msg_no_report": "Please upload data first to generate a report.",
        "msg_saved": "Saved",
        "col_issue_type": "Issue Type",
        "col_rows": "Rows",
        "sheet_summary": "Summary",
        "report_missing_uid": "Missing UID (Raw)",
        "report_cancelled": "Cancelled/deleted events (Raw)",
        "report_multi_version": "Multiple versions, same UID (Raw)",
        "report_duplicate_event_key": "Exact EVENT_KEY duplicates (Raw)",
        "report_same_patient_time": "Same patient & time (Clean, suspected)",
        "report_none": "No obvious duplicates / anomalies detected",
        # Visit type editor
        "btn_settings": "⚙️ Visit Types",
        "rules_dialog_title": "Visit Types & Prices",
        "rules_help": "Add the visit types you use. When importing a calendar, keywords are matched top-down.",
        "rules_col_code": "Code",
        "rules_col_keywords": "Match keywords (comma-separated)",
        "rules_col_amount": "Amount ($)",
        "rules_col_oasis": "OASIS required",
        "btn_rule_add": "+ Add",
        "btn_rule_del": "- Delete",
        "btn_rule_up": "↑ Move up",
        "btn_rule_down": "↓ Move down",
        "btn_rule_reset": "Reset to defaults",
        "msg_rule_invalid": "Code cannot be empty (row {row})",
        "msg_rule_duplicate": "Duplicate code: \"{code}\"",
        "msg_rule_bad_amount": "Amount must be a number (row {row})",
    },
}


class T:
    lang = "zh"

    @staticmethod
    def tr(key: str, **kwargs) -> str:
        s = TRANSLATIONS.get(T.lang, TRANSLATIONS["zh"]).get(key, key)
        if kwargs:
            try:
                s = s.format(**kwargs)
            except Exception:
                pass
        return s


tr = T.tr


# 报告类别 → 翻译 key（顺序固定）
REPORT_KEYS = [
    "report_missing_uid",
    "report_cancelled",
    "report_multi_version",
    "report_duplicate_event_key",
    "report_same_patient_time",
]


# =====================
# 访视类型规则（可由用户配置）
# =====================

# 默认规则——首次启动时用，用户可在 UI 里改
DEFAULT_RULES = [
    {"code": "SOC",    "keywords": ["SOC"],                      "amount": 100, "oasis": True},
    {"code": "RECERT", "keywords": ["RECERT"],                   "amount": 70,  "oasis": True},
    {"code": "ROC",    "keywords": ["ROC"],                      "amount": 70,  "oasis": True},
    {"code": "DC",     "keywords": ["DC"],                       "amount": 60,  "oasis": True},
    {"code": "FU",     "keywords": ["FOLLOW UP", "F/U", "FU"],   "amount": 65,  "oasis": False},
    {"code": "IV",     "keywords": ["IV"],                       "amount": 65,  "oasis": False},
]

YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NO_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
CLEAN_COLS = ["Date", "Time In", "Time Out", "Patient Name", "Visit Type", "Location", "Amount", "OASIS Required?"]


def load_rules() -> list[dict]:
    """从 QSettings 加载用户自定义规则；没有就返回默认值。"""
    settings = QSettings("nurseapp", "ics-visits-manager")
    raw = settings.value("rules_json", "", type=str)
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, list) and all(isinstance(r, dict) for r in data):
                # 兜底字段
                cleaned = []
                for r in data:
                    cleaned.append({
                        "code":     str(r.get("code", "")).strip(),
                        "keywords": [str(k).strip() for k in r.get("keywords", []) if str(k).strip()],
                        "amount":   int(r.get("amount", 0) or 0),
                        "oasis":    bool(r.get("oasis", False)),
                    })
                # 过滤掉 code 为空的
                cleaned = [r for r in cleaned if r["code"]]
                if cleaned:
                    return cleaned
        except Exception:
            pass
    return [dict(r) for r in DEFAULT_RULES]


def save_rules(rules: list[dict]):
    settings = QSettings("nurseapp", "ics-visits-manager")
    settings.setValue("rules_json", json.dumps(rules, ensure_ascii=False))


def apply_rules(raw_type: str | None, rules: list[dict]) -> tuple[str | None, int, str]:
    """
    用规则匹配 raw_type，返回 (标准化代号, 金额, OASIS 是否必需)
    规则按从上到下的顺序检查，第一条命中的规则生效。
    没匹配上：保留原始字符串，金额 0，OASIS NO。
    """
    if not raw_type:
        return None, 0, "NO"
    s = raw_type.upper().replace("-", " ").strip()
    for rule in rules:
        for kw in rule.get("keywords", []):
            if kw and kw.upper() in s:
                return rule["code"], int(rule.get("amount", 0)), ("YES" if rule.get("oasis") else "NO")
    return s, 0, "NO"


# =====================
# 时间安全处理
# =====================

def to_date_time(dt):
    if dt is None:
        return None, None
    if isinstance(dt, datetime):
        return dt.date().isoformat(), dt.time().strftime("%H:%M:%S")
    if isinstance(dt, date):
        return dt.isoformat(), None
    return str(dt), None

def to_iso(dt):
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.isoformat(sep=" ")
    if isinstance(dt, date):
        return dt.isoformat()
    return str(dt)


# =====================
# Summary 拆解
# =====================

def split_summary(summary: str | None):
    if not summary:
        return None, None
    s = summary.replace("—", "-").replace("–", "-").strip()
    if " - " in s:
        left, right = s.split(" - ", 1)
        return left.strip(), right.strip()
    if "-" in s:
        left, right = s.split("-", 1)
        return left.strip(), right.strip()
    return s.strip(), None


# =====================
# 取消/删除事件识别
# =====================

def is_cancelled_event(ev) -> bool:
    status = ev.get("status")
    if status and str(status).upper() == "CANCELLED":
        return True
    summary = ev.get("summary")
    if summary:
        s = str(summary).strip().lower()
        if s.startswith("cancelled") or s.startswith("canceled") or s.startswith("cancel"):
            return True
    return False

def calendar_method_is_cancel(cal: Calendar) -> bool:
    m = cal.get("method")
    return bool(m and str(m).upper() == "CANCEL")


# =====================
# ICS 解析
# =====================

def parse_ics(
    file_bytes: bytes,
    include_cancelled_in_clean: bool = False,
    rules: list[dict] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if rules is None:
        rules = DEFAULT_RULES
    cal = Calendar.from_ical(file_bytes)
    rows_clean, rows_raw = [], []
    cal_cancel_mode = calendar_method_is_cancel(cal)

    def safe_int(x, default=0):
        try: return int(x)
        except: return default

    for ev in cal.walk("VEVENT"):
        summary     = str(ev.get("summary"))     if ev.get("summary")     else None
        location    = str(ev.get("location"))    if ev.get("location")    else None
        uid         = str(ev.get("uid"))          if ev.get("uid")         else None
        description = str(ev.get("description")) if ev.get("description") else None
        status      = str(ev.get("status")).upper() if ev.get("status")   else None
        sequence    = safe_int(ev.get("sequence"), 0)

        last_modified  = ev.get("last-modified").dt  if ev.get("last-modified")  else None
        dtstamp        = ev.get("dtstamp").dt         if ev.get("dtstamp")        else None
        recurrence_id  = ev.get("recurrence-id").dt   if ev.get("recurrence-id") else None
        start          = ev.get("dtstart").dt          if ev.get("dtstart")       else None
        end            = ev.get("dtend").dt            if ev.get("dtend")         else None

        start_iso = to_iso(start)
        end_iso   = to_iso(end)
        d, tin    = to_date_time(start)
        _, tout   = to_date_time(end)

        patient, raw_type = split_summary(summary)
        vt_norm, vt_amount, vt_oasis = apply_rules(raw_type, rules)
        cancelled = cal_cancel_mode or is_cancelled_event(ev) or (status == "CANCELLED")

        uid_fallback = uid if uid else f"NOUID|{summary or ''}|{location or ''}"
        rid_iso      = to_iso(recurrence_id) if recurrence_id else ""
        event_key    = f"{uid_fallback}|RID={rid_iso}|{start_iso}|{end_iso}|{summary or ''}|{location or ''}"

        rows_raw.append({
            "EVENT_KEY": event_key, "UID": uid,
            "RECURRENCE-ID": to_iso(recurrence_id) if recurrence_id else None,
            "SEQUENCE": sequence, "LAST-MODIFIED": to_iso(last_modified),
            "DTSTAMP": to_iso(dtstamp), "STATUS": status,
            "CANCELLED?": "YES" if cancelled else "NO",
            "Summary (Raw)": summary, "Patient Name (Parsed)": patient,
            "Visit Type (Raw)": raw_type, "DTSTART (Raw)": start_iso,
            "DTEND (Raw)": end_iso, "Date (Parsed)": d,
            "Time In (Parsed)": tin, "Time Out (Parsed)": tout,
            "Location (Raw)": location, "Description (Raw)": description,
        })

        if cancelled and not include_cancelled_in_clean:
            continue

        rows_clean.append({
            "EVENT_KEY": event_key, "Date": d, "Time In": tin, "Time Out": tout,
            "Patient Name": patient, "Visit Type": vt_norm, "Location": location,
            "Amount": vt_amount, "OASIS Required?": vt_oasis,
        })

    df_raw   = pd.DataFrame(rows_raw)
    df_clean = pd.DataFrame(rows_clean)

    if len(df_raw):
        df_raw["_lm"] = pd.to_datetime(df_raw["LAST-MODIFIED"], errors="coerce")
        df_raw["_ds"] = pd.to_datetime(df_raw["DTSTAMP"],        errors="coerce")
        df_raw = df_raw.sort_values(
            by=["UID", "RECURRENCE-ID", "DTSTART (Raw)", "SEQUENCE", "_lm", "_ds"],
            na_position="first",
        )
        has_uid  = df_raw["UID"].notna() & (df_raw["UID"].astype(str).str.strip() != "")
        non_empty = df_raw[has_uid].copy()
        empty_uid = df_raw[~has_uid].copy()
        if len(non_empty):
            non_empty = non_empty.drop_duplicates(
                subset=["UID", "RECURRENCE-ID", "DTSTART (Raw)"], keep="last"
            )
        df_raw = pd.concat([non_empty, empty_uid], ignore_index=True)
        df_raw = df_raw.drop_duplicates(subset=["EVENT_KEY"], keep="last")
        df_raw = df_raw.drop(columns=["_lm", "_ds"], errors="ignore")

        keep_keys = set(df_raw["EVENT_KEY"].dropna())
        if len(df_clean):
            df_clean = df_clean[df_clean["EVENT_KEY"].isin(keep_keys)].copy()

    return df_raw, df_clean


# =====================
# Excel 导出
# =====================

def export_excel_clean(df: pd.DataFrame, path: Path):
    cols = [c for c in CLEAN_COLS if c in df.columns]
    df[cols].to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    oc = cols.index("OASIS Required?") + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, oc)
        cell.fill = YES_FILL if cell.value == "YES" else NO_FILL
    wb.save(path)

def export_excel_raw(df_raw: pd.DataFrame, path: Path):
    df_raw.to_excel(path, index=False)

def export_excel_bundle(df_clean: pd.DataFrame, df_raw: pd.DataFrame, path: Path):
    cols = [c for c in CLEAN_COLS if c in df_clean.columns]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_clean[cols].to_excel(writer, sheet_name="Clean_Amount", index=False)
        df_raw.to_excel(writer, sheet_name="Raw_Original", index=False)
        ws = writer.book["Clean_Amount"]
        header = [c.value for c in ws[1]]
        oc = header.index("OASIS Required?") + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, oc)
            cell.fill = YES_FILL if cell.value == "YES" else NO_FILL


# =====================
# 重复检测报告（用稳定 key，显示时翻译）
# =====================

def build_duplicate_report(df_raw: pd.DataFrame, df_clean: pd.DataFrame) -> dict:
    tables: dict[str, pd.DataFrame] = {}  # key 为翻译 key

    if len(df_raw):
        miss = df_raw[df_raw["UID"].isna() | (df_raw["UID"].astype(str).str.strip() == "")]
        if len(miss):
            tables["report_missing_uid"] = miss.copy()

    if len(df_raw) and "CANCELLED?" in df_raw.columns:
        canc = df_raw[df_raw["CANCELLED?"] == "YES"]
        if len(canc):
            tables["report_cancelled"] = canc.copy()

    if len(df_raw):
        key_cols = ["UID", "RECURRENCE-ID", "DTSTART (Raw)"]
        tmp = df_raw[df_raw["UID"].notna() & (df_raw["UID"].astype(str).str.strip() != "")].copy()
        if len(tmp):
            vc = tmp.groupby(key_cols, dropna=False).size().reset_index(name="count")
            multi = vc[vc["count"] > 1]
            if len(multi):
                mv = tmp.merge(multi[key_cols], on=key_cols, how="inner")
                tables["report_multi_version"] = mv.sort_values(key_cols + ["SEQUENCE"], na_position="last")

    if len(df_raw) and "EVENT_KEY" in df_raw.columns:
        dk = df_raw[df_raw.duplicated(subset=["EVENT_KEY"], keep=False)].copy()
        if len(dk):
            tables["report_duplicate_event_key"] = dk

    if len(df_clean):
        tmpc = df_clean.copy()
        tmpc["Date"] = pd.to_datetime(tmpc["Date"], errors="coerce").dt.date
        base = ["Date", "Time In", "Patient Name"]
        if all(c in tmpc.columns for c in base):
            ct = tmpc.groupby(base, dropna=False).size().reset_index(name="count")
            ct = ct[ct["count"] > 1]
            if len(ct):
                tables["report_same_patient_time"] = (
                    tmpc.merge(ct[base], on=base, how="inner").sort_values(base)
                )

    return {"tables": tables}


def report_summary_df(report: dict) -> pd.DataFrame:
    """根据当前语言生成汇总表"""
    tables = report["tables"]
    issue_col = tr("col_issue_type")
    rows_col = tr("col_rows")
    if not tables:
        return pd.DataFrame([{issue_col: tr("report_none"), rows_col: 0}])
    return pd.DataFrame([
        {issue_col: tr(k), rows_col: int(len(v))} for k, v in tables.items()
    ])


def export_dup_report_excel(report: dict, path: Path):
    def safe(name: str) -> str:
        for b in ['\\', '/', '*', '?', ':', '[', ']']:
            name = name.replace(b, "_")
        return name[:31]

    summary = report_summary_df(report)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name=tr("sheet_summary")[:31], index=False)
        for k, df_ in report["tables"].items():
            sn = safe(tr(k))
            if df_ is None or df_.empty:
                pd.DataFrame([{"info": "no rows"}]).to_excel(writer, sheet_name=sn, index=False)
            else:
                df_.to_excel(writer, sheet_name=sn, index=False)


# =====================
# 访视类型编辑对话框
# =====================

class RuleEditorDialog(QDialog):
    """让用户编辑访视类型 / 关键词 / 金额 / OASIS"""

    def __init__(self, rules: list[dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle(tr("rules_dialog_title"))
        self.resize(720, 480)
        self._rules_result: list[dict] | None = None
        self._build_ui()
        self._load(rules)

    def _build_ui(self):
        v = QVBoxLayout(self)
        v.setSpacing(8)

        v.addWidget(QLabel(tr("rules_help")))

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels([
            tr("rules_col_code"),
            tr("rules_col_keywords"),
            tr("rules_col_amount"),
            tr("rules_col_oasis"),
        ])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.verticalHeader().setVisible(True)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        v.addWidget(self.table)

        # 操作按钮
        bh = QHBoxLayout()
        b_add = QPushButton(tr("btn_rule_add"))
        b_del = QPushButton(tr("btn_rule_del"))
        b_up  = QPushButton(tr("btn_rule_up"))
        b_dn  = QPushButton(tr("btn_rule_down"))
        b_rst = QPushButton(tr("btn_rule_reset"))
        b_add.clicked.connect(self._add_row)
        b_del.clicked.connect(self._del_row)
        b_up.clicked.connect(lambda: self._move_row(-1))
        b_dn.clicked.connect(lambda: self._move_row(+1))
        b_rst.clicked.connect(self._reset_defaults)
        bh.addWidget(b_add)
        bh.addWidget(b_del)
        bh.addWidget(b_up)
        bh.addWidget(b_dn)
        bh.addStretch()
        bh.addWidget(b_rst)
        v.addLayout(bh)

        # 保存 / 取消
        bb = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        bb.accepted.connect(self._on_save)
        bb.rejected.connect(self.reject)
        v.addWidget(bb)

    def _load(self, rules: list[dict]):
        self.table.setRowCount(0)
        for r in rules:
            self._append_row(r)

    def _append_row(self, rule: dict):
        i = self.table.rowCount()
        self.table.insertRow(i)
        self.table.setItem(i, 0, QTableWidgetItem(rule.get("code", "")))
        self.table.setItem(i, 1, QTableWidgetItem(", ".join(rule.get("keywords", []))))
        self.table.setItem(i, 2, QTableWidgetItem(str(rule.get("amount", 0))))
        chk = QTableWidgetItem()
        chk.setFlags(chk.flags() | Qt.ItemFlag.ItemIsUserCheckable)
        chk.setCheckState(Qt.CheckState.Checked if rule.get("oasis") else Qt.CheckState.Unchecked)
        chk.setText("")
        self.table.setItem(i, 3, chk)

    def _add_row(self):
        self._append_row({"code": "", "keywords": [], "amount": 0, "oasis": False})
        self.table.selectRow(self.table.rowCount() - 1)

    def _del_row(self):
        i = self.table.currentRow()
        if i >= 0:
            self.table.removeRow(i)

    def _move_row(self, delta: int):
        i = self.table.currentRow()
        if i < 0:
            return
        j = i + delta
        if j < 0 or j >= self.table.rowCount():
            return
        # 交换两行内容
        for col in range(self.table.columnCount()):
            a = self.table.takeItem(i, col)
            b = self.table.takeItem(j, col)
            self.table.setItem(i, col, b)
            self.table.setItem(j, col, a)
        self.table.selectRow(j)

    def _reset_defaults(self):
        self._load([dict(r) for r in DEFAULT_RULES])

    def _on_save(self):
        out: list[dict] = []
        seen_codes: set[str] = set()
        for i in range(self.table.rowCount()):
            code = (self.table.item(i, 0).text() if self.table.item(i, 0) else "").strip()
            kw_text = (self.table.item(i, 1).text() if self.table.item(i, 1) else "").strip()
            amt_text = (self.table.item(i, 2).text() if self.table.item(i, 2) else "0").strip() or "0"
            chk = self.table.item(i, 3)
            oasis = (chk.checkState() == Qt.CheckState.Checked) if chk else False

            if not code:
                QMessageBox.warning(self, tr("rules_dialog_title"), tr("msg_rule_invalid", row=i + 1))
                return
            if code in seen_codes:
                QMessageBox.warning(self, tr("rules_dialog_title"), tr("msg_rule_duplicate", code=code))
                return
            seen_codes.add(code)
            try:
                amount = int(float(amt_text))
            except ValueError:
                QMessageBox.warning(self, tr("rules_dialog_title"), tr("msg_rule_bad_amount", row=i + 1))
                return

            keywords = [k.strip() for k in kw_text.split(",") if k.strip()]
            if not keywords:
                # 关键词为空时默认用代号本身做匹配
                keywords = [code]

            out.append({
                "code": code,
                "keywords": keywords,
                "amount": amount,
                "oasis": oasis,
            })

        self._rules_result = out
        self.accept()

    def result_rules(self) -> list[dict] | None:
        return self._rules_result


# =====================
# PyQt6 UI helpers
# =====================

def make_table() -> QTableWidget:
    t = QTableWidget()
    t.setAlternatingRowColors(True)
    t.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
    t.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
    t.verticalHeader().setVisible(False)
    t.setSortingEnabled(True)
    t.horizontalHeader().setStretchLastSection(True)
    return t

def fill_table(table: QTableWidget, df: pd.DataFrame, oasis_col: str | None = None):
    table.setSortingEnabled(False)
    table.clearContents()
    table.setRowCount(0)

    if df is None or df.empty:
        table.setColumnCount(0)
        return

    cols = df.columns.tolist()
    table.setColumnCount(len(cols))
    table.setHorizontalHeaderLabels(cols)
    table.setRowCount(len(df))

    for row_idx, (_, row) in enumerate(df.iterrows()):
        for col_idx, col_name in enumerate(cols):
            val  = row[col_name]
            text = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val)
            item = QTableWidgetItem(text)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

            if oasis_col and col_name == oasis_col:
                if text == "YES":
                    item.setBackground(QColor("#C6EFCE"))
                    item.setForeground(QColor("#276221"))
                elif text == "NO":
                    item.setBackground(QColor("#FFC7CE"))
                    item.setForeground(QColor("#9C0006"))

            table.setItem(row_idx, col_idx, item)

    table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
    table.horizontalHeader().setStretchLastSection(True)
    table.setSortingEnabled(True)


# =====================
# Main window
# =====================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # 加载保存的语言
        self.settings = QSettings("nurseapp", "ics-visits-manager")
        T.lang = self.settings.value("language", "zh", type=str)
        if T.lang not in TRANSLATIONS:
            T.lang = "zh"

        self.resize(1400, 900)

        self.df_raw      = pd.DataFrame()
        self.df_clean    = pd.DataFrame()
        self.df_filt     = pd.DataFrame()
        self.df_filt_raw = pd.DataFrame()
        self._report: dict | None = None
        self._loaded_filename: str = ""
        self._last_ics_bytes: bytes | None = None

        # 加载用户自定义的访视类型规则
        self._rules: list[dict] = load_rules()

        self._build_ui()
        self._retranslate_ui()

    # ------------------------------------------------------------------ build

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)

        # toolbar
        bar = QHBoxLayout()
        self.btn_upload = QPushButton()
        self.btn_upload.setFixedHeight(34)
        self.btn_upload.clicked.connect(self.upload_ics)

        self.btn_reset = QPushButton()
        self.btn_reset.setFixedHeight(34)
        self.btn_reset.clicked.connect(self.reset_data)

        self.btn_settings = QPushButton()
        self.btn_settings.setFixedHeight(34)
        self.btn_settings.clicked.connect(self.open_rules_editor)

        self.chk_cancelled = QCheckBox()

        self.lbl_lang = QLabel()
        self.lang_combo = QComboBox()
        self.lang_combo.addItem("中文", "zh")
        self.lang_combo.addItem("English", "en")
        self.lang_combo.setCurrentIndex(0 if T.lang == "zh" else 1)
        self.lang_combo.currentIndexChanged.connect(self._on_lang_changed)

        bar.addWidget(self.btn_upload)
        bar.addWidget(self.btn_reset)
        bar.addWidget(self.btn_settings)
        bar.addSpacing(16)
        bar.addWidget(self.chk_cancelled)
        bar.addStretch()
        bar.addWidget(self.lbl_lang)
        bar.addWidget(self.lang_combo)
        layout.addLayout(bar)

        # tabs
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)
        self._build_clean_tab()
        self._build_raw_tab()
        self._build_dup_tab()

    def _build_clean_tab(self):
        w = QWidget()
        v = QVBoxLayout(w)
        v.setSpacing(6)

        # date filter
        self.grp_date = QGroupBox()
        h = QHBoxLayout(self.grp_date)
        h.setSpacing(8)

        self.lbl_quick = QLabel()
        h.addWidget(self.lbl_quick)
        self.preset_combo = QComboBox()
        # 用 6 个空 item 占位，retranslate 时填充
        for _ in range(6):
            self.preset_combo.addItem("")
        self.preset_combo.setCurrentIndex(1)
        self.preset_combo.currentIndexChanged.connect(self._on_preset_changed)
        h.addWidget(self.preset_combo)
        h.addSpacing(12)

        self.lbl_from = QLabel()
        h.addWidget(self.lbl_from)
        self.date_start = QDateEdit(QDate.currentDate())
        self.date_start.setCalendarPopup(True)
        self.date_start.setEnabled(False)
        self.date_start.dateChanged.connect(self._apply_filter)
        h.addWidget(self.date_start)

        self.lbl_to = QLabel()
        h.addWidget(self.lbl_to)
        self.date_end = QDateEdit(QDate.currentDate())
        self.date_end.setCalendarPopup(True)
        self.date_end.setEnabled(False)
        self.date_end.dateChanged.connect(self._apply_filter)
        h.addWidget(self.date_end)
        h.addStretch()
        v.addWidget(self.grp_date)

        # metrics
        mh = QHBoxLayout()
        self.lbl_amount = QLabel()
        self.lbl_visits = QLabel()
        bold = QFont()
        bold.setPointSize(13)
        bold.setBold(True)
        self.lbl_amount.setFont(bold)
        self.lbl_visits.setFont(bold)
        mh.addWidget(self.lbl_amount)
        mh.addSpacing(40)
        mh.addWidget(self.lbl_visits)
        mh.addStretch()
        v.addLayout(mh)

        # table
        self.clean_table = make_table()
        v.addWidget(self.clean_table)

        # export
        eh = QHBoxLayout()
        self.btn_export_clean = QPushButton()
        self.btn_export_clean.clicked.connect(self.export_clean)
        self.btn_export_bundle = QPushButton()
        self.btn_export_bundle.clicked.connect(self.export_bundle)
        eh.addWidget(self.btn_export_clean)
        eh.addWidget(self.btn_export_bundle)
        eh.addStretch()
        v.addLayout(eh)

        self.tab_clean_widget = w
        self.tabs.addTab(w, "")

    def _build_raw_tab(self):
        w = QWidget()
        v = QVBoxLayout(w)

        self.raw_table = make_table()
        v.addWidget(self.raw_table)

        eh = QHBoxLayout()
        self.btn_export_raw = QPushButton()
        self.btn_export_raw.clicked.connect(self.export_raw)
        eh.addWidget(self.btn_export_raw)
        eh.addStretch()
        v.addLayout(eh)

        self.tab_raw_widget = w
        self.tabs.addTab(w, "")

    def _build_dup_tab(self):
        w = QWidget()
        v = QVBoxLayout(w)
        v.setSpacing(6)

        self.lbl_dup_summary = QLabel()
        v.addWidget(self.lbl_dup_summary)
        self.dup_summary = make_table()
        self.dup_summary.setMaximumHeight(120)
        v.addWidget(self.dup_summary)

        self.lbl_dup_detail = QLabel()
        v.addWidget(self.lbl_dup_detail)
        self.dup_detail_tabs = QTabWidget()
        v.addWidget(self.dup_detail_tabs)

        eh = QHBoxLayout()
        self.btn_export_dup = QPushButton()
        self.btn_export_dup.clicked.connect(self.export_dup)
        eh.addWidget(self.btn_export_dup)
        eh.addStretch()
        v.addLayout(eh)

        self.tab_dup_widget = w
        self.tabs.addTab(w, "")

    # ------------------------------------------------------------------ i18n

    def _retranslate_ui(self):
        self.setWindowTitle(tr("window_title"))

        self.btn_upload.setText(tr("btn_upload"))
        self.btn_reset.setText(tr("btn_reset"))
        self.btn_settings.setText(tr("btn_settings"))
        self.chk_cancelled.setText(tr("chk_cancelled"))
        self.lbl_lang.setText(tr("lang_label"))

        self.tabs.setTabText(0, tr("tab_clean"))
        self.tabs.setTabText(1, tr("tab_raw"))
        self.tabs.setTabText(2, tr("tab_dup"))

        # date filter
        self.grp_date.setTitle(tr("grp_date"))
        self.lbl_quick.setText(tr("lbl_quick"))
        self.lbl_from.setText(tr("lbl_from"))
        self.lbl_to.setText(tr("lbl_to"))

        # 不触发 currentIndexChanged
        self.preset_combo.blockSignals(True)
        preset_keys = ["preset_today", "preset_week", "preset_month",
                       "preset_last_month", "preset_all", "preset_custom"]
        for i, k in enumerate(preset_keys):
            self.preset_combo.setItemText(i, tr(k))
        self.preset_combo.blockSignals(False)

        # 导出按钮
        self.btn_export_clean.setText(tr("btn_export_clean"))
        self.btn_export_bundle.setText(tr("btn_export_bundle"))
        self.btn_export_raw.setText(tr("btn_export_raw"))
        self.btn_export_dup.setText(tr("btn_export_dup"))

        # 重复检测页
        self.lbl_dup_summary.setText(tr("lbl_dup_summary"))
        self.lbl_dup_detail.setText(tr("lbl_dup_detail"))

        # 数据相关：金额、状态栏、重复表格
        if self.df_filt is not None and not self.df_filt.empty:
            total = int(self.df_filt["Amount"].sum()) if "Amount" in self.df_filt.columns else 0
            self.lbl_amount.setText(tr("lbl_amount", amount=f"{total:,}"))
            self.lbl_visits.setText(tr("lbl_visits", count=len(self.df_filt)))
        else:
            self.lbl_amount.setText(tr("lbl_amount", amount="0"))
            self.lbl_visits.setText(tr("lbl_visits", count=0))

        # 状态栏
        if self._loaded_filename:
            self.statusBar().showMessage(tr(
                "status_loaded",
                name=self._loaded_filename,
                raw=len(self.df_raw),
                clean=len(self.df_clean),
            ))
        else:
            self.statusBar().showMessage(tr("status_no_data"))

        # 重复检测的汇总表 + 子标签
        if self._report is not None:
            fill_table(self.dup_summary, report_summary_df(self._report))
            self._rebuild_dup_detail_tabs()

    def _rebuild_dup_detail_tabs(self):
        self.dup_detail_tabs.clear()
        if self._report is None:
            return
        for k, tdf in self._report["tables"].items():
            t = make_table()
            fill_table(t, tdf)
            label = tr(k)
            label = label[:24] + ("…" if len(label) > 24 else "")
            self.dup_detail_tabs.addTab(t, label)

    def _on_lang_changed(self, idx: int):
        new_lang = self.lang_combo.itemData(idx) or "zh"
        if new_lang == T.lang:
            return
        T.lang = new_lang
        self.settings.setValue("language", new_lang)
        self._retranslate_ui()

    # ------------------------------------------------------------------ slots

    def upload_ics(self):
        path, _ = QFileDialog.getOpenFileName(
            self, tr("dlg_select_ics"), "", tr("dlg_ics_filter")
        )
        if not path:
            return
        try:
            with open(path, "rb") as f:
                data = f.read()
            self.df_raw, self.df_clean = parse_ics(
                data,
                include_cancelled_in_clean=self.chk_cancelled.isChecked(),
                rules=self._rules,
            )
        except Exception as e:
            QMessageBox.critical(self, tr("msg_parse_failed"), str(e))
            return
        self._last_ics_bytes = data
        self._loaded_filename = Path(path).name
        self._update_all()

    def open_rules_editor(self):
        dlg = RuleEditorDialog(self._rules, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_rules = dlg.result_rules()
            if new_rules is None:
                return
            self._rules = new_rules
            save_rules(new_rules)
            # 已有数据时用新规则重新解析
            if self._last_ics_bytes:
                try:
                    self.df_raw, self.df_clean = parse_ics(
                        self._last_ics_bytes,
                        include_cancelled_in_clean=self.chk_cancelled.isChecked(),
                        rules=self._rules,
                    )
                    self._update_all()
                except Exception as e:
                    QMessageBox.critical(self, tr("msg_parse_failed"), str(e))

    def reset_data(self):
        self.df_raw = self.df_clean = self.df_filt = self.df_filt_raw = pd.DataFrame()
        self._report = None
        self._loaded_filename = ""
        self._last_ics_bytes = None
        for t in (self.clean_table, self.raw_table, self.dup_summary):
            t.clearContents()
            t.setRowCount(0)
            t.setColumnCount(0)
        self.dup_detail_tabs.clear()
        self.lbl_amount.setText(tr("lbl_amount", amount="0"))
        self.lbl_visits.setText(tr("lbl_visits", count=0))
        self.statusBar().showMessage(tr("status_reset"))

    def _on_preset_changed(self):
        # 用 index 5 = custom，避免依赖语言
        is_custom = self.preset_combo.currentIndex() == 5
        self.date_start.setEnabled(is_custom)
        self.date_end.setEnabled(is_custom)
        self._apply_filter()

    def _apply_filter(self):
        if self.df_clean.empty:
            return

        df = self.df_clean.copy()
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
        valid = df["Date"].dropna()
        if valid.empty:
            return

        today = date.today()
        min_d, max_d = valid.min(), valid.max()
        idx = self.preset_combo.currentIndex()

        if idx == 0:        # today
            start = end = today
        elif idx == 1:      # this week
            start = today - timedelta(days=today.weekday())
            end = today
        elif idx == 2:      # this month
            start = today.replace(day=1)
            end = today
        elif idx == 3:      # last month
            first_this = today.replace(day=1)
            end = first_this - timedelta(days=1)
            start = end.replace(day=1)
        elif idx == 4:      # all
            start, end = min_d, max_d
        else:               # custom
            qs = self.date_start.date()
            qe = self.date_end.date()
            start = date(qs.year(), qs.month(), qs.day())
            end   = date(qe.year(), qe.month(), qe.day())

        self.df_filt = df[(df["Date"] >= start) & (df["Date"] <= end)].copy()

        raw = self.df_raw.copy()
        if not raw.empty and "Date (Parsed)" in raw.columns:
            raw["Date (Parsed)"] = pd.to_datetime(raw["Date (Parsed)"], errors="coerce").dt.date
            self.df_filt_raw = raw[
                (raw["Date (Parsed)"] >= start) & (raw["Date (Parsed)"] <= end)
            ].copy()
        else:
            self.df_filt_raw = raw

        total = int(self.df_filt["Amount"].sum()) if "Amount" in self.df_filt.columns else 0
        self.lbl_amount.setText(tr("lbl_amount", amount=f"{total:,}"))
        self.lbl_visits.setText(tr("lbl_visits", count=len(self.df_filt)))

        cols = [c for c in CLEAN_COLS if c in self.df_filt.columns]
        sorted_filt = self.df_filt.sort_values(["Date", "Time In"], na_position="last")
        fill_table(self.clean_table, sorted_filt[cols], oasis_col="OASIS Required?")

    def _update_all(self):
        fill_table(self.raw_table, self.df_raw)

        if not self.df_clean.empty:
            df = self.df_clean.copy()
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
            valid = df["Date"].dropna()
            if not valid.empty:
                mn, mx = valid.min(), valid.max()
                self.date_start.setMinimumDate(QDate(mn.year, mn.month, mn.day))
                self.date_start.setMaximumDate(QDate(mx.year, mx.month, mx.day))
                self.date_end.setMinimumDate(QDate(mn.year, mn.month, mn.day))
                self.date_end.setMaximumDate(QDate(mx.year, mx.month, mx.day))
                self.date_end.setDate(QDate(mx.year, mx.month, mx.day))

        self._apply_filter()

        if not self.df_raw.empty or not self.df_clean.empty:
            self._report = build_duplicate_report(self.df_raw, self.df_clean)
            fill_table(self.dup_summary, report_summary_df(self._report))
            self._rebuild_dup_detail_tabs()

        self.statusBar().showMessage(tr(
            "status_loaded",
            name=self._loaded_filename,
            raw=len(self.df_raw),
            clean=len(self.df_clean),
        ))

    # ------------------------------------------------------------------ export

    def _save_path(self, title: str, default: str) -> Path | None:
        p, _ = QFileDialog.getSaveFileName(self, title, default, tr("dlg_xlsx_filter"))
        return Path(p) if p else None

    def export_clean(self):
        if self.df_filt.empty:
            return QMessageBox.warning(self, tr("msg_no_data_title"), tr("msg_no_data_filt"))
        p = self._save_path(tr("dlg_save_clean"), "HH_Visits_With_Amount.xlsx")
        if p:
            export_excel_clean(self.df_filt, p)
            QMessageBox.information(self, tr("msg_saved"), str(p))

    def export_raw(self):
        if self.df_raw.empty:
            return QMessageBox.warning(self, tr("msg_no_data_title"), tr("msg_no_data_raw"))
        p = self._save_path(tr("dlg_save_raw"), "HH_Visits_Raw.xlsx")
        if p:
            raw = self.df_filt_raw if not self.df_filt_raw.empty else self.df_raw
            export_excel_raw(raw, p)
            QMessageBox.information(self, tr("msg_saved"), str(p))

    def export_bundle(self):
        if self.df_filt.empty:
            return QMessageBox.warning(self, tr("msg_no_data_title"), tr("msg_no_data_filt"))
        p = self._save_path(tr("dlg_save_bundle"), "HH_Visits_Bundle.xlsx")
        if p:
            raw = self.df_filt_raw if not self.df_filt_raw.empty else self.df_raw
            export_excel_bundle(self.df_filt, raw, p)
            QMessageBox.information(self, tr("msg_saved"), str(p))

    def export_dup(self):
        if self._report is None:
            return QMessageBox.warning(self, tr("msg_no_report_title"), tr("msg_no_report"))
        p = self._save_path(tr("dlg_save_dup"), "Duplicate_Report.xlsx")
        if p:
            export_dup_report_excel(self._report, p)
            QMessageBox.information(self, tr("msg_saved"), str(p))


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
